import asyncio
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment
from io import BytesIO

def snake_to_text(text):
    return text.replace('_', ' ').replace('-', ' ').capitalize()


def create_xlsx(header: list, mandatory_fields: list):
    wb = Workbook()
    ws = wb.active

    # formated_mandatory_fields = [
    #     snake_to_text(key) for key in mandatory_fields
    # ]

    formated_header = []
    ''' adding (*) to the imported keys'''
    for index, key in enumerate(header):
        if key in mandatory_fields:
            print(key+"*")
            formated_header.append(snake_to_text(key+'*'))
        else:
            formated_header.append(snake_to_text(key))

    print(formated_header)
    ws.append(formated_header)

    font = Font(size=14, bold=True, color='FFFFFFFF')
    fill = PatternFill(start_color="FF13D133", end_color="FF13D133", fill_type='solid')
    # red_fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type='solid')
    align = Alignment(horizontal='center', vertical='center')

    for cell in ws[1]:
        cell.font = font
        cell.alignment = align
        cell.fill = fill

    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return stream

def read_xlsx(file):
    wb = load_workbook(file)
    ws = wb.active

    rows = ws.iter_rows(values_only=True)
    headers = next(rows)

    '''formating the header'''
    header = [header.replace(" ","_").lower().rstrip("*") for header in headers]
    for row in rows:
        yield dict(zip(header, row))

def write_xlsx(file: Workbook, row_object: dict):
    wb = load_workbook(file)
    ws = wb.active

    '''we have to consider that the worksheet that has been uploaded is the correct worksheet for the typeid
        and contains the necessary headers
    '''

    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    '''formating the header'''

    header = [header.replace(" ", "_").lower() for header in headers]
    try:
        '''match the header with the dict and make a new list of values mathcing the list of headers'''
        row = [row_object.get(key) for key in header]

        ws.append(row)
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
    except:
        return "error"
    return buffer

def remove_row(file: Workbook, index: int):
    wb = load_workbook(file)
    ws = wb.active

    ws.delete_rows(index)
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer

if __name__ == "__main__":
    asyncio.to_thread(create_xlsx(["col1", "last_visit"]))
